import imaplib
import email
from email.header import decode_header
import openpyxl  # Para manejar Excel
import os
import re  # Para buscar el número en el asunto

# Función para limpiar etiquetas HTML con expresiones regulares
def limpiar_html(texto):
    # Elimina cualquier etiqueta HTML
    texto_limpio = re.sub(r'<.*?>', '', texto)
    return texto_limpio

import html  # Importa el módulo html

def extraer_datos_correo(body_limpio, consecutivo):
    # Decodifica las entidades HTML
    body_limpio = html.unescape(body_limpio)
    # Expresiones regulares para cada campo
    fecha = re.search(r"Fecha:\s*(.+)", body_limpio)
    
    razon_social = re.search(r"Razón Social:\s*(.+)", body_limpio)
    rfc = re.search(r"RFC:\s*(.+)", body_limpio)
    sucursal = re.search(r"Sucursal:\s*(.+)", body_limpio)
    regimen_fiscal = re.search(r"Régimen Fiscal:\s*(.+)", body_limpio)
    uso_cfdi = re.search(r"Uso de CFDI:\s*(.+)", body_limpio)
    forma_pago = re.search(r"Forma de Pago:\s*(.+)", body_limpio)
    codigo_postal = re.search(r"Código Postal:\s*(.+)", body_limpio)
    correo_electronico = re.search(r"Correo Electrónico:\s*(.+)", body_limpio)
    folio = re.search(r"Folio:\s*(.+)", body_limpio)
    monto = re.search(r"Monto:\s*(.+)", body_limpio)

    # Convertir los valores a texto si se encuentran
    datos = [
        fecha.group(1).strip() if fecha else "No encontrada",
        razon_social.group(1).strip() if razon_social else "No encontrada",
        rfc.group(1).strip() if rfc else "No encontrado",
        sucursal.group(1).strip() if sucursal else "No encontrada",
        regimen_fiscal.group(1).strip() if regimen_fiscal else "No encontrado",
        uso_cfdi.group(1).strip() if uso_cfdi else "No encontrado",
        forma_pago.group(1).strip() if forma_pago else "No encontrada",
        codigo_postal.group(1).strip() if codigo_postal else "No encontrado",
        correo_electronico.group(1).strip() if correo_electronico else "No encontrado",
        folio.group(1).strip() if folio else "No encontrado",
        monto.group(1).strip() if monto else "No encontrado",
        consecutivo
    ]
    return datos

# Función para guardar la información extraída en datos_factura.xlsx
def guardar_en_excel_datos_factura(datos):
    archivo_excel = "datos_factura.xlsx"  # Archivo para datos de factura
    
    # Revisar si el archivo existe
    if os.path.exists(archivo_excel):
        wb = openpyxl.load_workbook(archivo_excel)
        print("Archivo datos_factura.xlsx encontrado y cargado.")
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Agregar encabezados si es un archivo nuevo
        encabezados = [
            "Fecha", "Razón Social", "RFC", "Sucursal", "Régimen Fiscal", 
            "Uso de CFDI", "Forma de Pago", "Código Postal", 
            "Correo Electrónico", "Folio", "Monto", "Consecutivo"
        ]
        sheet.append(encabezados)
        print("Archivo datos_factura.xlsx creado con encabezados.")

    # Seleccionar la hoja activa
    sheet = wb.active
    # Guardar los datos en una nueva fila
    sheet.append(datos)
    
    # Guardar el archivo Excel
    wb.save(archivo_excel)
    wb.close()
    print("Datos guardados correctamente en datos_factura.xlsx.")
    
    # Mover el archivo a la carpeta TEMPORAL
    mover_a_temporal(archivo_excel)

# Función para mover archivo a la carpeta TEMPORAL
def mover_a_temporal(archivo_excel):
    ruta_destino = "../TEMPORAL/"  # Ruta de la carpeta TEMPORAL
    if os.path.exists(archivo_excel):
        nuevo_nombre = os.path.join(ruta_destino, os.path.basename(archivo_excel))
        os.rename(archivo_excel, nuevo_nombre)
        print(f"Archivo Excel movido a la carpeta: {ruta_destino}")
    else:
        print("El archivo datos_factura.xlsx no existe o ya fue movido.")

# Función para guardar la información extraída en Log.xlsx
def guardar_en_log(datos):
    archivo_excel = "../Log/Log.xlsx"  # Ruta al archivo Log.xlsx
    
    # Revisar si el archivo existe
    if os.path.exists(archivo_excel):
        wb = openpyxl.load_workbook(archivo_excel)
        print("Archivo Log.xlsx encontrado y cargado.")
    else:
        print("Error: El archivo Log.xlsx no existe en la ruta especificada.")
        return

    # Seleccionar la hoja activa
    sheet = wb.active
    
    # Encontrar la primera fila vacía
    primera_fila_vacia = None
    for row in range(2, sheet.max_row + 2):  # Comenzamos en la fila 2 para evitar encabezados
        if not any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            primera_fila_vacia = row
            break

    # Si no se encuentra una fila vacía, se asume que la siguiente fila disponible es la correcta
    if primera_fila_vacia is None:
        primera_fila_vacia = sheet.max_row + 1

    # Guardar los datos en la primera fila vacía encontrada
    for col, dato in enumerate(datos, start=1):
        sheet.cell(row=primera_fila_vacia, column=col, value=dato)
    
    # Guardar el archivo Excel
    wb.save(archivo_excel)
    wb.close()
    print(f"Datos guardados correctamente en la fila {primera_fila_vacia} del archivo Log.xlsx.")

# Crear conexión con el servidor de correo
imap = imaplib.IMAP4_SSL("imap.gmail.com")
# Iniciar sesión
imap.login('gallito2.empresys@gmail.com', 'ygvt fmgi clfj uccw')

# Seleccionar bandeja de entrada y buscar solo correos no leídos
imap.select("INBOX")
status, mensajes_no_leidos = imap.search(None, 'UNSEEN')

# Obtener la lista de IDs de los correos no leídos
mensajes_no_leidos = mensajes_no_leidos[0].split()

if mensajes_no_leidos:
    # Leer el último correo no leído
    ultimo_mensaje_id = mensajes_no_leidos[-1]
    res, mensaje = imap.fetch(ultimo_mensaje_id, "(RFC822)")
    
    for respuesta in mensaje:
        if isinstance(respuesta, tuple):
            # Obtener el mensaje de correo
            msg = email.message_from_bytes(respuesta[1])
            subject = decode_header(msg["Subject"])[0][0]
            if isinstance(subject, bytes):
                subject = subject.decode(errors='ignore')  # Ignorar errores de decodificación

            # Verificar si el asunto contiene "Solicitud de factura - " seguido de un número
            match = re.search(r"Solicitud de factura - (\d+)", subject)
            if match:
                consecutivo = match.group(1)
                print(f"Consecutivo encontrado: {consecutivo}")

                # Si el correo tiene múltiples partes (por ejemplo, texto y HTML)
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        content_disposition = str(part.get("Content-Disposition"))
                        try:
                            body = part.get_payload(decode=True).decode()
                        except:
                            pass
                        if content_type == "text/html" and "attachment" not in content_disposition:
                            # Limpiar las etiquetas HTML
                            body_limpio = limpiar_html(body)
                            print(body_limpio)
                            
                            # Extraer los datos específicos del cuerpo del correo usando regex
                            datos = extraer_datos_correo(body_limpio, consecutivo)
                            
                            # Guardar los datos en los dos archivos
                            guardar_en_excel_datos_factura(datos)
                            guardar_en_log(datos)
                            break
                else:
                    # Si el correo no tiene múltiples partes, limpiar el HTML
                    body = msg.get_payload(decode=True).decode()
                    body_limpio = limpiar_html(body)
                    
                    # Extraer los datos específicos del cuerpo del correo usando regex
                    datos = extraer_datos_correo(body_limpio, consecutivo)

                    # Guardar los datos en los dos archivos
                    guardar_en_excel_datos_factura(datos)
                    guardar_en_log(datos)

                # Marcar el correo como leído
                imap.store(ultimo_mensaje_id, '+FLAGS', '\\Seen')
else:
    print("No hay correos no leídos.")

# Cerrar la conexión con el servidor
imap.close()
imap.logout()
