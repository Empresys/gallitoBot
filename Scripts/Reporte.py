import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import pandas as pd
from datetime import datetime, date, time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta para el archivo de registro de envíos
registro_envios = "../Folder/registro_envios.txt"

# Definir las ventanas de tiempo permitidas como una variable global
ventanas = [
    (time(9, 0), time(9, 45)),
    (time(13, 0), time(13, 45)),
    (time(18, 0), time(18, 45))
]

# Función para verificar si la hora actual está dentro de las ventanas de tiempo permitidas
def en_hora_permitida():
    hora_actual = datetime.now().time()
    
    # Verificar si la hora actual está dentro de alguna de las ventanas
    for inicio, fin in ventanas:
        if inicio <= hora_actual <= fin:
            print(f"Hora actual {hora_actual} está dentro de la ventana {inicio}-{fin}.")
            return True
    print(f"Hora actual {hora_actual} no está en ninguna ventana permitida.")
    return False

# Función para verificar si ya se ha registrado un envío hoy dentro de la ventana actual
def envio_ya_registrado():
    fecha_hoy = date.today().isoformat()
    hora_actual = datetime.now().time()
    registros = []

    # Leer el archivo de registro si existe
    if os.path.exists(registro_envios):
        with open(registro_envios, "r") as file:
            registros = file.readlines()

    # Verificar si ya se ha enviado en el rango horario actual
    for registro in registros:
        # Ignorar líneas vacías o mal formateadas
        if not registro.strip():
            continue
        
        try:
            registro_fecha, registro_hora = registro.strip().split()
            registro_hora = datetime.strptime(registro_hora, "%H:%M").time()
            
            # Revisar si el envío fue hoy y está dentro de la misma ventana horaria
            if registro_fecha == fecha_hoy:
                for inicio, fin in ventanas:
                    if inicio <= registro_hora <= fin and inicio <= hora_actual <= fin:
                        print(f"Ya se ha registrado un envío hoy dentro de la ventana horaria {inicio}-{fin}.")
                        return True
        except ValueError:
            print(f"Línea en el archivo de registro con formato incorrecto: {registro.strip()}")

    return False

# Función para registrar un envío de hoy
def registrar_envio():
    fecha_hoy = date.today().isoformat()
    hora_actual = datetime.now().strftime("%H:%M")
    with open(registro_envios, "a") as file:
        file.write(f"{fecha_hoy} {hora_actual}\n")

# Función para resaltar columnas específicas en el archivo Excel temporal
def resaltar_columnas(archivo):
    wb = load_workbook(archivo)
    ws = wb.active

    # Resaltar la columna 'Folio' (anteriormente 'Consecutivo') en amarillo
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in ws['F']:  # Asumiendo que 'Folio' ahora es la primera columna en Log_temp.xlsx
        cell.fill = yellow_fill

    # Resaltar 'Finalizado' en rojo para 'ERROR' y verde para 'EXITOSO'
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
        for cell in row:
            if cell.value == "ERROR":
                cell.fill = red_fill
            elif cell.value == "Exitoso":
                cell.fill = green_fill

    # Guardar los cambios
    wb.save(archivo)

# Función para limpiar el archivo original 'Log.xlsx' manteniendo las cabeceras
def limpiar_archivo_log(archivo_log):
    wb = load_workbook(archivo_log)
    ws = wb.active

    # Mantener solo la primera fila (cabecera) y eliminar el resto
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None  # Limpiar cada celda en las filas que se eliminan

    wb.save(archivo_log)
    print("Archivo original 'Log.xlsx' ha sido limpiado, manteniendo las cabeceras.")

# Función para enviar el correo
def enviar_log_por_correo(archivo_log):
    if not en_hora_permitida() or envio_ya_registrado():
        print("No es la hora permitida o ya se ha enviado el correo en esta ventana.")
        return

    # Leer el archivo de Excel y eliminar las columnas no deseadas
    df = pd.read_excel(archivo_log)
    columnas_a_eliminar = ['RFC', 'Régimen Fiscal', 'Uso de CFDI', 'Forma de Pago', 'Código Postal', 'Folio', 'Intentos']
    df_modificado = df.drop(columns=[col for col in columnas_a_eliminar if col in df.columns])

    # Renombrar la columna 'Consecutivo' a 'Folio'
    if 'Consecutivo' in df_modificado.columns:
        df_modificado = df_modificado.rename(columns={'Consecutivo': 'Folio'})
    
    # Guardar el archivo temporal sin columnas eliminadas
    archivo_temporal = "Log_temp.xlsx"
    df_modificado.to_excel(archivo_temporal, index=False)
    
    # Resaltar las columnas en el archivo temporal
    resaltar_columnas(archivo_temporal)

    # Configuración del correo
    remitente = 'gallito2.empresys@gmail.com'
    destinatario = 'leomonroy2017@gmail.com'
    copia = 'seuperez@empresys.com'
    fecha_hora_actual = datetime.now().strftime("%Y-%m-%d %H:%M")
    asunto = f'Log del proceso - {fecha_hora_actual}'
    
    cuerpo_mensaje = 'Buen día, se adjunta el log del proceso'
    
    # Crear el mensaje de correo
    msg = MIMEMultipart()
    msg['From'] = remitente
    msg['To'] = destinatario
    msg['Cc'] = copia
    msg['Subject'] = asunto
    msg.attach(MIMEText(cuerpo_mensaje, 'plain'))
    
    # Adjuntar el archivo temporal
    with open(archivo_temporal, 'rb') as adjunto:
        parte = MIMEBase('application', 'octet-stream')
        parte.set_payload(adjunto.read())
        encoders.encode_base64(parte)
        parte.add_header('Content-Disposition', f"attachment; filename={os.path.basename(archivo_temporal)}")
        msg.attach(parte)
    
    # Conectar con el servidor SMTP y enviar el correo
    try:
        servidor_smtp = smtplib.SMTP('smtp.gmail.com', 587)
        servidor_smtp.starttls()
        servidor_smtp.login(remitente, 'ygvt fmgi clfj uccw')  # Usa la contraseña de aplicación
        texto = msg.as_string()
        servidor_smtp.sendmail(remitente, [destinatario] + copia.split(','), texto)
        servidor_smtp.quit()
        print(f"Correo enviado exitosamente a {destinatario} con copia a {copia}.")
        
        # Registrar el envío
        registrar_envio()
        
        # Limpiar el archivo original 'Log.xlsx'
        limpiar_archivo_log(archivo_log)
        
        # Eliminar el archivo temporal
        os.remove(archivo_temporal)
        
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

# Llamada a la función con la ruta del archivo Log.xlsx
archivo_log = "../Log/Log.xlsx"
enviar_log_por_correo(archivo_log)
